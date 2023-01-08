from openpyxl import load_workbook

Kinds = []
Kinds_num = []
Product = []
pcode = []

Kinds_num_list_num = -1
Kinds_num_num = 1
sheet_num = 1

wb = load_workbook('Product.xlsx')
ws = wb.active

for i in range(2, 55):
  if ws.cell(i,1).value not in Kinds:
    Kinds.append(ws.cell(i,1).value)
    Kinds_num_list_num += 1
    Kinds_num_num = 1
    sheet_num = 1
    Kinds_num.append(Kinds_num_num)
    wb.create_sheet('{}'.format(ws.cell(i,1).value))
    ws_create = wb['{}'.format(wb.sheetnames[-1])]
    print(ws_create)
    for x in range(2, 4):
      ws_create.cell(sheet_num, x-1).value = ws.cell(i, x).value
  else:
    Kinds_num[Kinds_num_list_num] += 1
    sheet_num += 1
    for x in range(2, 4):
      ws_create.cell(sheet_num, x-1).value = ws.cell(i, x).value
  

print(Kinds)
print(Kinds_num)
print(Product)
print(pcode)

print(ws)
print()
print(ws_create)
# print(ws.cell(i,3).value)

wb.save('Product.xlsx')